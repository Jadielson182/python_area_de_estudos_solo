# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
PASTA_XLSX = ROOT_FOLDER / 'XLSX'
WORKBOOK_PATH = PASTA_XLSX / 'workbook.xlsx'

PASTA_XLSX.mkdir(exist_ok=True)

workbook = Workbook()
# esse abaixo e pra usar planilha padrao ativa
# worksheet: Worksheet = workbook.active  

# aqui foi criada nome da planilha com variavel
sheet_name = 'Minha Planilha'

#aqui foi criada a planilha, OBS: o 0 representa indice, indica que a planilha e o indice 0 que vai ser a primeira a ser selecionada
workbook.create_sheet(sheet_name, 0)

#Aqui selecionou a planilha, OBS: ou pode usar planilha ativa usando workbook.active ou botar o nome da planilha em [] como exemplo abaixo,
# mas nesse caso ja foi selecionado em cima usando indice no exemplo acima
worksheet: Worksheet = workbook[sheet_name]

#Remover uma planilha
# workbook.remove(workbook['Sheet'])


#criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'nota')
worksheet.cell(1, 4, 'Media')


students = [
    #Nome   idade  nota
    ['João',  14,   5.5, 10],
    ['Maria', 13,   9.2, 8],
    ['Luiz',  15,   8.8, 9],
    ['Alberto', 16, 10, 6],
    ['Carlos', 17, 9.5, 8],
    ['Tereza ', 15, 16, 5 ],
    ['Marcos', 17, 18, 8],
    ['Diana', 21, 19, 9 ],
    ['Paulo', 20, 17, 6 ],
    ['Debora', 16, 15, 9 ],
    ['Patricia', 18, 19, 8],

]
#logica mais simples
# for student in students:
#     worksheet.append(student)

#Logica mais complexa, porem que deve lembrar

for i, students_row in enumerate(students, start=2):
    for j, students_column in enumerate(students_row, start=1):
        worksheet.cell(i, j, students_column)



workbook.save(WORKBOOK_PATH)
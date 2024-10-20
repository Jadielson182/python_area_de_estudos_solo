# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell  # esse foi para a tipagem
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
PASTA_XLSX = ROOT_FOLDER / 'XLSX'
WORKBOOK_PATH = PASTA_XLSX / 'workbook.xlsx'

PASTA_XLSX.mkdir(exist_ok=True)

#carregando arquivo do excel
workbook = load_workbook(WORKBOOK_PATH)


# aqui foi criada nome da planilha com variavel
sheet_name = 'Minha Planilha'

#Aqui selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]  # isso e uma tipagem pq o vscode pode da alguns erro de tipagem
for row in worksheet.iter_rows(min_row=2, min_col=1):
    for col in row:
        print(col.value, end='\t')
        if Cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23) # no () cell.row e indicar linha e 2 indica a coluna e 23 o valor que sera adicionado
    print()

# aqui é uma maneira de altera dados usando indice da planilha
worksheet['B3'].value = 14  


workbook.save(WORKBOOK_PATH)